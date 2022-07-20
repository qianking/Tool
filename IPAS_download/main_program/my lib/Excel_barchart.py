from turtle import title
import openpyxl
from copy import deepcopy
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference

Site= '蘇州'
#User_select_project = "SWITCH_CISCO_EZ1KA1"
Model = 'all'
Line = "all"
Device = "all"
#Time_period = '2022/05/19 20:00:00 ~ 2022/05/20 08:00:00'
#today_excute_download_path = 'D:\Qian\python\Tool\畫excel'
#Station_Data= {0: {'Pretest': '7/356', 'Retest_psss': {'Loopback-Test-SFP-1G-Link-TEST (QSFC09)': ['2222013600430', '2226913100148', '2226913100115'], 'Boot-Up (QSFCG4)': ['2226913100171'], 'MCU-bootloader-version (QSFF02)': ['2222013601940'], 'SFIS-Get-Timer (Q0FT18)': ['2222013600274'], 'I2C-SFP1 (Q0FT18)': ['2222013600442']}, 'Fail': {}}, 1: {'Hi_pot_test': '0/0'}, 2: {'Final_Test': '14/683', 'Retest_psss': {'MOUNT-USB-Check (QSFI02)': ['2226913101395', '2222013600409', '2222013600912', '2222013600868'], 'BurnIn-Check (QSFT15)': ['2226913100866', '2222013601981', '2226913100267'], 'SFP-Link-Check (QSFC31)': ['2222013601936'], 'ControlBoard-Hook-Power-ON (QSFT27)': ['2222013600155'], 'Port-0-6-LinkUP (QSFC35)': ['2226913100009']}, 'Fail': {'PG-Test (QSFT0P)': ['2226913101169']}}, 3: {'Final_Check': '52/727', 'Retest_psss': {'Get-ECC-SUDI (QSFT40)': ['2222013600234', '2226913101092', '2226913100515'], 'CLIIP-install (QSFT55)': ['2222013601325', '2222013601069', '2226913100326', '2222013601676', '2222013600133', '2226913100144', '2226913100328'], 'DUT-Turn-On (QSFT15)': ['2222013601384', '2222013601679', '2226913101644', '2222013600747', '2222013601372', '2222013600944'], 'Get-RSA-SUDI (QSFT40)': ['2222013601749', '2226913100135', '2226913100577', '2222013601762'], 'Get-AIK-SUDI (QSFT40)': ['2226913100012', '2222013601622', '2226913100408'], 'DUT-Timeout (Q0FT18)': ['2226913100286', '2226913100116'], 'RJ-Green-LED-Turn-on (QSFL06)': ['2226913100939'], 'System-Info-Check (QSFT15)': ['2222013601094'], 'ECC-SUDI-Install (QSFT60)': ['2226913100907'], 'PCAMAP-MAC-Addr-Check (QSFM01)': ['2226913100044'], 'Verification (QSFT61)': ['2222013600859']}, 
#'Fail': {'Get-ECC-SUDI (QSFT40)': ['2222013600775', '2222013600272', '2222013601234', '2222013601123'], 'CLIIP-install (QSFT55)': ['2222013600374', '2222013600191'], 'Reading-SUDI-install (QSFT61)': ['2222013600708'], 'Get-AIK-SUDI (QSFT40)': ['2222013600141']}}, 4: {'Software_Download': '23/1184', 'Retest_psss': {'DUT-Bootloader-Reset-TIME-OUT (QSFCG4)': ['2226913100717', '2226913100817', '2222013601099', '2222013600895', '2226913100157', '2222013600282', '2222013600228', '2222013600301'], 'Boot-Loader-Check-Version (QSFF69)': ['2222013601460', '2222013600107', '2222013601394'], 'IOS-Boot-Up (QSFCG4)': ['2222013600290'], 'DUT-Reload-TIME-OUT (QSFCG4)': ['2226913100543', '2222013600417'], 'Boot-Loader-Check-Version-TIME-OUT (Q0FT18)': ['2226913100136']}, 'Fail': {'DUT-Bootloader-Reset-TIME-OUT (QSFCG4)': ['2222013600152'], 'Boot-Loader-Check-Version (QSFF69)': ['2222013601091']}}}


def return_to_excel_data(Station_Data):
    
    excel_data = []
    temp_list = []
    temp_station_list = []
    all_stations = []
    pass_isn_num =[]
    pass_isn = []

    fail_isn_num = []
    fail_isn = []
    sum_num = []
    for ind in range(len(Station_Data)):
        del temp_list[:]
        del all_stations[:]
        del pass_isn_num [:]
        del pass_isn[:]

        del fail_isn_num [:]
        del fail_isn[:]
        del sum_num[:]

        Test_station = list(Station_Data[ind].keys())[0]
        fail_num = list(Station_Data[ind].values())[0].split('/')[0]
        
        if int(fail_num) != 0:
            pass_list = list(Station_Data[ind].values())[1]
            pass_stations = list(pass_list.keys())
            isns = list(pass_list.values())
            all_stations = deepcopy(pass_stations)
            for isn in isns:
                pass_isn_num.append(len(isn))
            pass_isn = deepcopy(isns)
            
            fail_list = list(Station_Data[ind].values())[2]
            fail_stations = list(fail_list.keys())
            fail_isns = list(fail_list.values())
            while(len(pass_isn_num) > len(fail_isn_num)):
                        fail_isn_num.append(0)
                        fail_isn.append('')
            for fail in range(len(fail_stations)):
                if fail_stations[fail] in all_stations:
                    index = all_stations.index(fail_stations[fail])              
                    fail_isn_num[index] = len(fail_isns[fail])
                    fail_isn[index] = fail_isns[fail]
                else:
                    all_stations.append(deepcopy(fail_stations[fail]))
                    pass_isn_num.append(0)
                    pass_isn.append('')
                    fail_isn_num.append(len(fail_isns[fail]))
                    fail_isn.append(fail_isns[fail])
            if len(fail_stations) == 0:
                for k in range(len(all_stations)):
                    fail_isn_num.append(0)
                    fail_isn.append('')
            for i in range(len(pass_isn_num)):
                sum_num.append(pass_isn_num[i] + fail_isn_num[i])

            for l in range(len(sum_num)-1):
                for k in range(l+1, len(sum_num)):
                    if sum_num[k] > sum_num[l]:
                        temp = sum_num[l]
                        sum_num[l] = sum_num[k]
                        sum_num[k] = temp
                        
                        temp = all_stations[l]
                        all_stations[l] = all_stations[k]
                        all_stations[k] = temp
                        
                        temp = pass_isn_num[l]
                        pass_isn_num[l] = pass_isn_num[k]
                        pass_isn_num[k] = temp
                        
                        temp = fail_isn_num[l]
                        fail_isn_num[l] = fail_isn_num[k]
                        fail_isn_num[k] = temp

                        temp = pass_isn[l]
                        pass_isn[l] = pass_isn[k]
                        pass_isn[k] = temp

                        temp = fail_isn[l]
                        fail_isn[l] = fail_isn[k]
                        fail_isn[k] = temp

            temp_list = [Test_station, all_stations, pass_isn_num, fail_isn_num, sum_num, pass_isn, fail_isn] 
            excel_data.append(deepcopy(temp_list))
    #print('\n')    
    #print(excel_data)
    #print('\n')  

    excel_data_flip = []

    for index in range(len(excel_data)):
        del temp_list[:]
        del temp_station_list[:]

        station = excel_data[index][0]
        temp_list.append(deepcopy(station))
        error_station = excel_data[index][1]
        pass_num = excel_data[index][2]
        fail_num = excel_data[index][3]
        sum_num = excel_data[index][4]
        pass_isn = excel_data[index][5]
        fail_isn = excel_data[index][6]
        for o in range(len(error_station)):
            error_name = error_station[o].split(' ')[0]
            error_code = error_station[o].split(' ')[1]
            error_code = error_code[1:-1]
            pass_isns = "、".join(pass_isn[o])
            fail_isns = "、".join(fail_isn[o])

            temp_station_list = [error_name, error_code, pass_num[o], fail_num[o], sum_num[o], pass_isns, fail_isns]
            temp_list.append(deepcopy(temp_station_list))
        excel_data_flip.append(deepcopy(temp_list))
    
    #print(excel_data_flip)
    return excel_data_flip

def draw_on_excel(Station_Data, User_select_project, Time_period, today_excute_download_path):
    information = [f'Site : {Site}', f'Project : {User_select_project}', f'Model : {Model}', f'Line : {Line}', f'Device : {Device}', f'Time : {Time_period}']
    save_path = f'{today_excute_download_path}\Fail_report_SZ_{User_select_project}.xlsx'
    book = Workbook()
    try:
        excel_data = return_to_excel_data(Station_Data)
    except:
        return 'Failed to create the fail report!'
    else:
        #region 設定第一頁資訊頁
        sheet = book.active  
        sheet.title = 'information'
        for g in range(1,7):
            sheet.column_dimensions["A"].width = 100
            cell = sheet.cell(row = g, column=1)
            cell.value = information[g-1]
            cell.font = Font(name = "Calibri", size = 12, bold = True)
            cell.alignment = Alignment(horizontal = 'center')
            cell.fill = PatternFill(fill_type="solid", fgColor= "FFE5CC")
            cell.border = Border(left=Side(style='thin', color='330000'), right=Side(style='thin', color='330000'),top=Side(style='thin', color='330000'), bottom=Side(style='thin', color='330000'))
        #endregion


        for i in range(len(excel_data)):
            sheet_name = excel_data[i][0]
            sheet_name =sheet_name.replace('_', ' ')
            sheet = book.create_sheet(sheet_name)
            sheet.column_dimensions["A"].width = 40
            sheet.column_dimensions["B"].width = 20
            sheet.column_dimensions["C"].width = 20
            sheet.column_dimensions["D"].width = 15
            sheet.column_dimensions["E"].width = 15
            sheet.column_dimensions["F"].width = 100
            sheet.column_dimensions["G"].width = 100
            sheet.merge_cells('A1:G1')
            cell = sheet.cell(row = 1, column=1)
            cell.value = f'Station : {sheet_name}'
            cell.fill = PatternFill(fill_type="solid", fgColor= "FFE5CC")
            cell.font = Font(name = "Calibri", size = 12, bold = True)
            cell.alignment = Alignment(horizontal = 'center')
            cell.border = Border(left=Side(style='thin', color='330000'), right=Side(style='thin', color='330000'),top=Side(style='thin', color='330000'), bottom=Side(style='thin', color='330000'))
            sheet.append(['Error Name', 'Error Code', 'Retest Pass Count', 'Fail Count', 'Total Count', 'Retest Pass Isn', 'Fail Isn'])
            for row in sheet.iter_rows(min_row=2, max_row=2, min_col=1):
                for cell in row:
                    cell.fill = PatternFill(fill_type="solid", fgColor= "00CCCC")
                    cell.font = Font(name = "Calibri", size = 12, bold = True)
                    cell.alignment = Alignment(horizontal = 'center')
                    cell.border = Border(left=Side(style='thin', color='330000'), right=Side(style='thin', color='330000'),top=Side(style='thin', color='330000'), bottom=Side(style='thin', color='330000'))

            for k in range(1, len(excel_data[i])):
                sheet.append(excel_data[i][k])
            
            for row in sheet.iter_rows(min_row=3, min_col=1, max_col=5):
                for cell in row:
                    cell.border = Border(left=Side(style='thin', color='330000'), right=Side(style='thin', color='330000'),top=Side(style='thin', color='330000'), bottom=Side(style='thin', color='330000'))  
                    cell.font = Font(name = "Calibri", size = 12)
                    cell.alignment = Alignment(horizontal = 'center')
            
            for row in sheet.iter_rows(min_row=3, min_col=6, max_col=7):
                for cell in row:
                    cell.border = Border(left=Side(style='thin', color='330000'), right=Side(style='thin', color='330000'),top=Side(style='thin', color='330000'), bottom=Side(style='thin', color='330000'))  
                    cell.font = Font(name = "Calibri", size = 12)
                    cell.alignment = Alignment(horizontal = 'left')
            

            if (len(excel_data[i])-1) >= 5:
                maxrow = 7
            else:
                maxrow = len(excel_data[i]) +1

            chart = BarChart()
            chart.type = 'bar'
            chart.style = 10
            chart.title = f'Top {maxrow-2}'
            chart.overlap = 100
            chart.grouping = 'stacked'
            data = Reference(sheet, min_col = 3, max_col = 4, min_row = 2, max_row = maxrow)
            cats = Reference(sheet, min_col = 1, min_row = 3, max_row = maxrow)
            chart.add_data(data, titles_from_data = True)
            chart.set_categories(cats)
            chart.shape = 4
            chart.y_axis.title = 'Error Count'
            chart.y_axis.majorUnit = 1
            chart.x_axis.scaling.orientation = "maxMin"
            chart.x_axis.tickLblPos = "low"
            chart.height = 10
            chart.width = 20
            sheet.add_chart(chart, f'A{len(excel_data[i])+4}')


        book.save(save_path)
        return 'Create the fail report successfully'

if '__main__' == __name__:
    pass
    #print(return_to_excel_data(Station_Data))
    #draw_on_excel(Station_Data, User_select_project, Time_period, today_excute_download_path)




   









        
            
        
        



