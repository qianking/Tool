from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Color
from openpyxl.styles import numbers

len_column = int()
len_row = int()
font = Font(name='標楷體', size=12)


def border_set(start_column, start_row , column, row):
    left = right = top = bottom = 'thin'
    
    if column == start_column:
        left = 'thick'
    if row == start_row:
        top = 'thick'
    if column == (start_column+len_column-1):
        right = 'thick'
    if row == (start_row+len_row-1):
        bottom = 'thick'
    
    return Border(left=Side(style=left), 
                     right=Side(style=right), 
                     top=Side(style=top), 
                     bottom=Side(style=bottom))


def Transfer_Excel(X_shear_data, Y_shear_data, excel_out):
    global len_column
    global len_row

    wb = Workbook()
    del wb['Sheet']

    len_column = len(X_shear_data[2])
    len_row = len(X_shear_data)

    floor_len = len(X_shear_data[3:])
    #floor_len = 14
        
    #創建資料頁面
    if floor_len > 15:  #假設樓層大於15樓就分兩頁
        wb.create_sheet('X報表')
        wb.create_sheet('Y報表')
        sheet_X = wb['X報表']
        sheet_Y = wb['Y報表']
    else:                           #假設樓層小於15樓就放在同一頁
        wb.create_sheet('XY報表')
        sheet_XY = wb['XY報表']

    #region x頁面
    start_row = 4
    start_column = 2 #起始位置

    if floor_len > 15:  #假設樓層大於15樓就分兩頁
        sheet = sheet_X
    else:
        sheet = sheet_XY

    cell = sheet.cell(row=2, column=1)
    cell.value = "極限層剪力"
    cell.font = Font(name='標楷體', size=12, bold=True)

    for i, x_shear in enumerate(X_shear_data, start=start_row):
        for j, x in enumerate(x_shear, start=start_column):
            cell = sheet.cell(row=i, column=j, value=x)
            
            if str(x).replace('.','',1).isdigit():
                cell.number_format = numbers.FORMAT_NUMBER_00

            cell.border = border_set(start_column, start_row, j,i)
            cell.font = font

            if x == "X向":
                cell.font = Font(name='標楷體', size=12, bold=True)
            if x == "NG":
                cell.font = Font(name='標楷體', size=12, color=Color(rgb="FFFFFF00"))
            

    rows, cols = sheet.max_row, sheet.max_column

    for i in range(1, rows+1):
        sheet.row_dimensions[i].height = 17

    # 設置列寬
    for i in range(1, cols+1):
        sheet.column_dimensions[sheet.cell(row=1, column=i).column_letter].width = 11
    #endregion


    #region y頁面
    if floor_len > 15:  #假設樓層大於15樓就分兩頁
        start_row = 2
        start_column = 2 #起始位置
        sheet = sheet_Y
    else:
        start_row = start_row + len_row + 2
        start_column = 2 #起始位置
        sheet = sheet_XY

    for i, y_shear in enumerate(Y_shear_data, start=start_row):
        for j, y in enumerate(y_shear, start=start_column):
            cell = sheet.cell(row=i, column=j, value=y)
            
            if str(y).replace('.','',1).isdigit():
                cell.number_format = numbers.FORMAT_NUMBER_00

            cell.border = border_set(start_column, start_row, j,i)
            cell.font = font

            if y == "Y向":
                cell.font = Font(name='標楷體', size=12, bold=True)
            if y == "NG":
                cell.font = Font(name='標楷體', size=12, color=Color(rgb="FFFFFF00"))
            

    rows, cols = sheet.max_row, sheet.max_column

    for i in range(1, rows+1):
        sheet.row_dimensions[i].height = 17

    # 設置列寬
    for i in range(1, cols+1):
        sheet.column_dimensions[sheet.cell(row=1, column=i).column_letter].width = 11
    
    #endregion
 
    wb.save(excel_out)



