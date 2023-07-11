from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles import numbers
import os
import win32com.client

len_column = int()
len_row = int()
start_row = 4
start_column = 2

def convert_to_number(s):
    # 檢查字串是否只包含數字
    if s.isdigit():        
        return int(s)
    # 檢查字串是否為浮點數
    elif s.replace('.', '', 1).isdigit() and s.count('.')<2:
        return float(s)  # 轉換成浮點數
    else:
        return s  # 不轉換

wb = Workbook()

# 選擇活動工作表
ws = wb.active
font = Font(name='標楷體', size=12)


def border_set(column, row):
    left = right = top = bottom = 'thin'
    
    if column == start_column:
        left = 'thick'
    if row == start_row:
        top = 'thick'
    if column == (start_column+len_column-1):
        right = 'thick'
    if row == (start_row+len_row-1):
        bottom = 'thick'
    
    return Border(left=Side(style=left), 
                     right=Side(style=right), 
                     top=Side(style=top), 
                     bottom=Side(style=bottom))


input_path = r'C:\Users\andy_chien\Downloads\弱\V600_弱層檢核\OUTPUT\VPDATXE.TXT'
out_path = r'C:\Users\andy_chien\Downloads\弱\V600_弱層檢核\OUTPUT\output.xlsx'
out_pdf_path = r'C:\Users\andy_chien\Downloads\弱\V600_弱層檢核\OUTPUT\output.pdf'

with open(input_path, 'r') as f:
    data = f.read()

data_tt = [a for a in data.strip().split("\n") if a != ""]
data_tt = [a.split() for a in data_tt]


len_column = len(data_tt[2])
len_row = len(data_tt)
for i in data_tt:
    for j in range(len_column-len(i)):
        i.append("")       
print(data_tt)



for i, a in enumerate(data_tt, start=start_row):
    for j, b in enumerate(a, start=start_column):
        c = convert_to_number(b)
        cell = ws.cell(row=i, column=j, value=c)

        if b.replace('.','',1).isdigit():
            cell.number_format = numbers.FORMAT_NUMBER_00

        cell.border = border_set(j,i)
        cell.font = font


rows, cols = ws.max_row, ws.max_column

for i in range(1, rows+1):
    ws.row_dimensions[i].height = 17

# 設置列寬
for i in range(1, cols+1):
    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 11

wb.save(out_path)
 


""" # 创建一个Excel应用对象
excel = win32com.client.Dispatch("Excel.Application")

# 打开Excel文件
wb = excel.Workbooks.Open(out_path)

# 访问活动工作表
ws = wb.Worksheets[0]  # 0代表第一个工作表

# 调整页面设置以适应所有列
ws.PageSetup.Zoom = False
ws.PageSetup.FitToPagesWide = 1  # 所有列适应在一个页面宽
ws.PageSetup.FitToPagesTall = False  # 页面的高度可以是任意的

# 导出为PDF
wb.ExportAsFixedFormat(0, out_pdf_path)

# 关闭Excel文件
wb.Close() """