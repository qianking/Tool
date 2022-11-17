import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import Reference, Series, BarChart
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties  

''' execute_data = {'site':'蘇州',
                'data_source':'Test Station',
                'user_select_project': 'SWITCH_CISCO_EZ1KA1',   
                'time_selection':{'time':'Select Manually', 'time_period':['2022/08/28 00:00', '2022/08/28 10:00']}} '''

execute_data = dict()
all_data = dict()
''' all_data = {"today_download_path": "D:\\IPLAS Download\\2022-09-07 10-09 SWITCH_CISCO_EZ1KA1", 
"station_data": {
		"Pretest": {"UPH": 0, "fail_num": 4, "pass_num": 273, 
				"ISN_data": {"System LED Check IS ON": {"error_code": "QSFL03", "error_count": "2", "fail_num": "1", "retest_pass_num": "0", "fail_list": {"2259346202095 08/28 05:02:25": "Pretest\\System_LED_Check_IS_ON\\2259346202095_995547_2022_08_28 05_02_25\\2259346202095_PRE-TEST(995547)_ON_C1000-8FP-2G-L_V1.01_NG(QSFL03).txt", "2259346202095 08/28 07:07:14": "Pretest\\System_LED_Check_IS_ON\\2259346202095_995548_2022_08_28 07_07_14\\2259346202095_PRE-TEST(995548)_ON_C1000-8FP-2G-L_V1.01_NG(QSFL03).txt"}, "retest_list": {}, "error_count_data": {"2259346202095 08/28 05:02:25": "Pretest\\System_LED_Check_IS_ON\\2259346202095_995547_2022_08_28 05_02_25\\2259346202095_PRE-TEST(995547)_ON_C1000-8FP-2G-L_V1.01_NG(QSFL03).txt", "2259346202095 08/28 07:07:14": "Pretest\\System_LED_Check_IS_ON\\2259346202095_995548_2022_08_28 07_07_14\\2259346202095_PRE-TEST(995548)_ON_C1000-8FP-2G-L_V1.01_NG(QSFL03).txt"}, "error_count_path": "Pretest\\System_LED_Check_IS_ON"}, 
							"Loopback Test-SFP 1G Link TEST": {"error_code": "QSFC09", "error_count": "1", "fail_num": "0", "retest_pass_num": "1", "fail_list": {}, "retest_list": {"2259346202141 08/28 02:10:52": "Pretest\\Loopback_Test_SFP_1G_Link_TEST\\2259346202141_995593_2022_08_28 02_10_52\\2259346202141_PRE-TEST(995593)_ON_C1000-8FP-2G-L_V1.01_NG(QSFC09).txt"}, "error_count_data": {"2259346202141 08/28 02:10:52": "Pretest\\Loopback_Test_SFP_1G_Link_TEST\\2259346202141_995593_2022_08_28 02_10_52\\2259346202141_PRE-TEST(995593)_ON_C1000-8FP-2G-L_V1.01_NG(QSFC09).txt"}, "error_count_path": "Pretest\\Loopback_Test_SFP_1G_Link_TEST"}, 
							"SFIS Get Timer": {"error_code": "Q0FT18", "error_count": "1", "fail_num": "0", "retest_pass_num": "1", "fail_list": {}, "retest_list": {"2259346202056 08/28 00:42:44": "Pretest\\SFIS_Get_Timer\\2259346202056_995593_2022_08_28 00_42_44\\2259346202056_PRE-TEST(995593)_ON_C1000-8FP-2G-L_V1.01_NG(Q0FT18).txt"}, "error_count_data": {"2259346202056 08/28 00:42:44": "Pretest\\SFIS_Get_Timer\\2259346202056_995593_2022_08_28 00_42_44\\2259346202056_PRE-TEST(995593)_ON_C1000-8FP-2G-L_V1.01_NG(Q0FT18).txt"}, "error_count_path": "Pretest\\SFIS_Get_Timer"}}, 
						"final_fail_num": 1, "final_pass_num": 2}, 
		
		"Hi_pot_test": {"UPH": 0, "fail_num": 0, "pass_num": 0, 
				"ISN_data": None, 
						"final_fail_num": 0, "final_pass_num": 0}, 
		
		"Final_Test": {"UPH": 0, "fail_num": 8, "pass_num": 233, 
				"ISN_data": {"Check Green LED Color": {"error_code": "QSFL06", "error_count": "2", "fail_num": "1", "retest_pass_num": "1", "fail_list": {"2259346200817 08/28 07:13:55": "Final Test\\Check_Green_LED_Color\\2259346200817_620886_2022_08_28 07_13_55\\PSZ26351XV9_FINALTEST(620886)_ON_C1000-8FP-2G-L_V1.01_NG(QSFL06).txt"}, "retest_list": {"2259346202989 08/28 09:39:53": "Final Test\\Check_Green_LED_Color\\2259346202989_620886_2022_08_28 09_39_53\\PSZ26351XKT_FINALTEST(620886)_ON_C1000-8FP-2G-L_V1.01_NG(QSFL06).txt"}, "error_count_data": {"2259346200817 08/28 07:13:55": "Final Test\\Check_Green_LED_Color\\2259346200817_620886_2022_08_28 07_13_55\\PSZ26351XV9_FINALTEST(620886)_ON_C1000-8FP-2G-L_V1.01_NG(QSFL06).txt", "2259346202989 08/28 09:39:53": "Final Test\\Check_Green_LED_Color\\2259346202989_620886_2022_08_28 09_39_53\\PSZ26351XKT_FINALTEST(620886)_ON_C1000-8FP-2G-L_V1.01_NG(QSFL06).txt"}, "error_count_path": "Final Test\\Check_Green_LED_Color"}, 
							"PG Test": {"error_code": "QSFT0P", "error_count": "2", "fail_num": "1", "retest_pass_num": "0", "fail_list": {"2260991901066 08/28 04:22:25": "Final Test\\PG_Test\\2260991901066_620893_2022_08_28 04_22_25\\PSZ26351XE1_FINALTEST(620893)_ON_C1000-8T-2G-L_V1.01_NG(QSFT0P).txt", "2260991901066 08/28 04:34:59": "Final Test\\PG_Test\\2260991901066_620886_2022_08_28 04_34_59\\PSZ26351XE1_FINALTEST(620886)_ON_C1000-8T-2G-L_V1.01_NG(QSFT0P).txt"}, "retest_list": {}, "error_count_data": {"2260991901066 08/28 04:22:25": "Final Test\\PG_Test\\2260991901066_620893_2022_08_28 04_22_25\\PSZ26351XE1_FINALTEST(620893)_ON_C1000-8T-2G-L_V1.01_NG(QSFT0P).txt", "2260991901066 08/28 04:34:59": "Final Test\\PG_Test\\2260991901066_620886_2022_08_28 04_34_59\\PSZ26351XE1_FINALTEST(620886)_ON_C1000-8T-2G-L_V1.01_NG(QSFT0P).txt"}, "error_count_path": "Final Test\\PG_Test"}, 
							"Port 0/6 LinkUP": {"error_code": "QSFC35", "error_count": "1", "fail_num": "0", "retest_pass_num": "1", "fail_list": {}, "retest_list": {"2259346202531 08/28 03:17:24": "Final Test\\Port_0_6_LinkUP\\2259346202531_620893_2022_08_28 03_17_24\\PSZ26351X20_FINALTEST(620893)_ON_C1000-8FP-2G-L_V1.01_NG(QSFC35).txt"}, "error_count_data": {"2259346202531 08/28 03:17:24": "Final Test\\Port_0_6_LinkUP\\2259346202531_620893_2022_08_28 03_17_24\\PSZ26351X20_FINALTEST(620893)_ON_C1000-8FP-2G-L_V1.01_NG(QSFC35).txt"}, "error_count_path": "Final Test\\Port_0_6_LinkUP"}, 
							"Port 0/2 LinkUP": {"error_code": "QSFC36", "error_count": "1", "fail_num": "0", "retest_pass_num": "1", "fail_list": {}, "retest_list": {"2259346201674 08/28 09:25:34": "Final Test\\Port_0_2_LinkUP\\2259346201674_620886_2022_08_28 09_25_34\\PSZ26351X99_FINALTEST(620886)_ON_C1000-8FP-2G-L_V1.01_NG(QSFC36).txt"}, "error_count_data": {"2259346201674 08/28 09:25:34": "Final Test\\Port_0_2_LinkUP\\2259346201674_620886_2022_08_28 09_25_34\\PSZ26351X99_FINALTEST(620886)_ON_C1000-8FP-2G-L_V1.01_NG(QSFC36).txt"}, "error_count_path": "Final Test\\Port_0_2_LinkUP"}, 
							"Port 0/4 LinkUP": {"error_code": "QSFC35", "error_count": "1", "fail_num": "0", "retest_pass_num": "0", "fail_list": {}, "retest_list": {}, "error_count_data": {"2259346200817 08/28 07:39:22": "Final Test\\Port_0_4_LinkUP\\2259346200817_620893_2022_08_28 07_39_22\\PSZ26351XV9_FINALTEST(620893)_ON_C1000-8FP-2G-L_V1.01_NG(QSFC35).txt"}, "error_count_path": "Final Test\\Port_0_4_LinkUP"}, 
							"Combo copper-prefer setup": {"error_code": "QSFC38", "error_count": "1", "fail_num": "0", "retest_pass_num": "1", "fail_list": {}, "retest_list": {"2259346200570 08/28 05:36:51": "Final Test\\Combo_copper_prefer_setup\\2259346200570_620886_2022_08_28 05_36_51\\PSZ26351Y9C_FINALTEST(620886)_ON_C1000-8FP-2G-L_V1.01_NG(QSFC38).txt"}, "error_count_data": {"2259346200570 08/28 05:36:51": "Final Test\\Combo_copper_prefer_setup\\2259346200570_620886_2022_08_28 05_36_51\\PSZ26351Y9C_FINALTEST(620886)_ON_C1000-8FP-2G-L_V1.01_NG(QSFC38).txt"}, "error_count_path": "Final Test\\Combo_copper_prefer_setup"}}, 
						"final_fail_num": 2, "final_pass_num": 4}, 
		
		"Final_Check": {"UPH": 0, "fail_num": 1, "pass_num": 265, 
				"ISN_data": {"RJ LED off": {"error_code": "QSFT15", "error_count": "1", "fail_num": "0", "retest_pass_num": "1", "fail_list": {}, "retest_list": {"2259346200823 08/28 05:46:53": "Final Check\\RJ_LED_off\\2259346200823_620834_2022_08_28 05_46_53\\PSZ26351XRQ_FINALCHECK(620834)_ON_C1000-8FP-2G-L_V1.01_NG(QSFT15).txt"}, "error_count_data": {"2259346200823 08/28 05:46:53": "Final Check\\RJ_LED_off\\2259346200823_620834_2022_08_28 05_46_53\\PSZ26351XRQ_FINALCHECK(620834)_ON_C1000-8FP-2G-L_V1.01_NG(QSFT15).txt"}, "error_count_path": "Final Check\\RJ_LED_off"}}, 
                        "final_fail_num": 0, "final_pass_num": 1}, 
		
        "Software_Download": {"UPH": 0, "fail_num": 15, "pass_num": 448, 
                "ISN_data": {"DUT Bootloader Reset TIME OUT": {"error_code": "QSFCG4", "error_count": "4", "fail_num": "0", "retest_pass_num": "4", "fail_list": {}, "retest_list": {"2259346202837 08/28 01:07:39": "Software Download\\DUT_Bootloader_Reset_TIME_OUT\\2259346202837_991033_2022_08_28 01_07_39\\PSZ26351X31_SOFTWAREDOWNLOAD(991033)_ON_C1000-8FP-2G-L_V1.01_NG(QSFCG4).txt", "2259346200191 08/28 02:51:54": "Software Download\\DUT_Bootloader_Reset_TIME_OUT\\2259346200191_991048_2022_08_28 02_51_54\\PSZ26351VEL_SOFTWAREDOWNLOAD(991048)_ON_C1000-8FP-2G-L_V1.01_NG(QSFCG4).txt", "2259346202114 08/28 06:48:37": "Software Download\\DUT_Bootloader_Reset_TIME_OUT\\2259346202114_991034_2022_08_28 06_48_37\\PSZ26351Y0U_SOFTWAREDOWNLOAD(991034)_ON_C1000-8FP-2G-L_V1.01_NG(QSFCG4).txt", "2259346201211 08/28 00:09:25": "Software Download\\DUT_Bootloader_Reset_TIME_OUT\\2259346201211_991036_2022_08_28 00_09_25\\PSZ26351UHJ_SOFTWAREDOWNLOAD(991036)_ON_C1000-8FP-2G-L_V1.01_NG(QSFCG4).txt"}, "error_count_data": {"2259346202837 08/28 01:07:39": "Software Download\\DUT_Bootloader_Reset_TIME_OUT\\2259346202837_991033_2022_08_28 01_07_39\\PSZ26351X31_SOFTWAREDOWNLOAD(991033)_ON_C1000-8FP-2G-L_V1.01_NG(QSFCG4).txt", "2259346200191 08/28 02:51:54": "Software Download\\DUT_Bootloader_Reset_TIME_OUT\\2259346200191_991048_2022_08_28 02_51_54\\PSZ26351VEL_SOFTWAREDOWNLOAD(991048)_ON_C1000-8FP-2G-L_V1.01_NG(QSFCG4).txt", "2259346202114 08/28 06:48:37": "Software Download\\DUT_Bootloader_Reset_TIME_OUT\\2259346202114_991034_2022_08_28 06_48_37\\PSZ26351Y0U_SOFTWAREDOWNLOAD(991034)_ON_C1000-8FP-2G-L_V1.01_NG(QSFCG4).txt", "2259346201211 08/28 00:09:25": "Software Download\\DUT_Bootloader_Reset_TIME_OUT\\2259346201211_991036_2022_08_28 00_09_25\\PSZ26351UHJ_SOFTWAREDOWNLOAD(991036)_ON_C1000-8FP-2G-L_V1.01_NG(QSFCG4).txt"}, "error_count_path": "Software Download\\DUT_Bootloader_Reset_TIME_OUT"}, 
							"IOS Boot Up": {"error_code": "QSFCG4", "error_count": "3", "fail_num": "0", "retest_pass_num": "2", "fail_list": {}, "retest_list": {"2259346200427 08/28 07:40:18": "Software Download\\IOS_Boot_Up\\2259346200427_991072_2022_08_28 07_40_18\\PSZ26351Y6E_SOFTWAREDOWNLOAD(991072)_ON_C1000-8FP-2G-L_V1.01_NG(QSFCG4).txt", "2259346201623 08/28 03:46:34": "Software Download\\IOS_Boot_Up\\2259346201623_991991_2022_08_28 03_46_34\\PSZ26351WJD_SOFTWAREDOWNLOAD(991991)_ON_C1000-8FP-2G-L_V1.01_NG(QSFCG4).txt"}, "error_count_data": {"2259346200427 08/28 07:40:18": "Software Download\\IOS_Boot_Up\\2259346200427_991072_2022_08_28 07_40_18\\PSZ26351Y6E_SOFTWAREDOWNLOAD(991072)_ON_C1000-8FP-2G-L_V1.01_NG(QSFCG4).txt", "2259346201623 08/28 03:46:34": "Software Download\\IOS_Boot_Up\\2259346201623_991991_2022_08_28 03_46_34\\PSZ26351WJD_SOFTWAREDOWNLOAD(991991)_ON_C1000-8FP-2G-L_V1.01_NG(QSFCG4).txt", "2259346201640 08/28 00:06:29": "Software Download\\IOS_Boot_Up\\2259346201640_991036_2022_08_28 00_06_29\\PSZ26351UX9_SOFTWAREDOWNLOAD(991036)_ON_C1000-8FP-2G-L_V1.01_NG(QSFCG4).txt"}, "error_count_path": "Software Download\\IOS_Boot_Up"}, 
							"IOS Ping Check": {"error_code": "QSFT15", "error_count": "2", "fail_num": "0", "retest_pass_num": "2", "fail_list": {}, "retest_list": {"2259346202769 08/28 00:24:04": "Software Download\\IOS_Ping_Check\\2259346202769_991991_2022_08_28 00_24_04\\PSZ26351V9X_SOFTWAREDOWNLOAD(991991)_ON_C1000-8FP-2G-L_V1.01_NG(QSFT15).txt", "2259346201703 08/28 00:20:55": "Software Download\\IOS_Ping_Check\\2259346201703_991951_2022_08_28 00_20_55\\PSZ26351WAW_SOFTWAREDOWNLOAD(991951)_ON_C1000-8FP-2G-L_V1.01_NG(QSFT15).txt"}, "error_count_data": {"2259346202769 08/28 00:24:04": "Software Download\\IOS_Ping_Check\\2259346202769_991991_2022_08_28 00_24_04\\PSZ26351V9X_SOFTWAREDOWNLOAD(991991)_ON_C1000-8FP-2G-L_V1.01_NG(QSFT15).txt", "2259346201703 08/28 00:20:55": "Software Download\\IOS_Ping_Check\\2259346201703_991951_2022_08_28 00_20_55\\PSZ26351WAW_SOFTWAREDOWNLOAD(991951)_ON_C1000-8FP-2G-L_V1.01_NG(QSFT15).txt"}, "error_count_path": "Software Download\\IOS_Ping_Check"}, 
							"TFTP Down IOS File": {"error_code": "QSFF23", "error_count": "1", "fail_num": "0", "retest_pass_num": "1", "fail_list": {}, "retest_list": {"2259346202537 08/28 05:49:16": "Software Download\\TFTP_Down_IOS_File\\2259346202537_991951_2022_08_28 05_49_16\\PSZ26351WJ0_SOFTWAREDOWNLOAD(991951)_ON_C1000-8FP-2G-L_V1.01_NG(QSFF23).txt"}, "error_count_data": {"2259346202537 08/28 05:49:16": "Software Download\\TFTP_Down_IOS_File\\2259346202537_991951_2022_08_28 05_49_16\\PSZ26351WJ0_SOFTWAREDOWNLOAD(991951)_ON_C1000-8FP-2G-L_V1.01_NG(QSFF23).txt"}, "error_count_path": "Software Download\\TFTP_Down_IOS_File"}, 
							"IOS Initial Dialog": {"error_code": "QSFT15", "error_count": "1", "fail_num": "0", "retest_pass_num": "1", "fail_list": {}, "retest_list": {"2259346202765 08/28 09:30:03": "Software Download\\IOS_Initial_Dialog\\2259346202765_991943_2022_08_28 09_30_03\\PSZ26351Y4R_SOFTWAREDOWNLOAD(991943)_ON_C1000-8FP-2G-L_V1.01_NG(QSFT15).txt"}, "error_count_data": {"2259346202765 08/28 09:30:03": "Software Download\\IOS_Initial_Dialog\\2259346202765_991943_2022_08_28 09_30_03\\PSZ26351Y4R_SOFTWAREDOWNLOAD(991943)_ON_C1000-8FP-2G-L_V1.01_NG(QSFT15).txt"}, "error_count_path": "Software Download\\IOS_Initial_Dialog"}, 
							"eeprom save": {"error_code": "QSFT15", "error_count": "1", "fail_num": "0", "retest_pass_num": "1", "fail_list": {}, "retest_list": {"2259346202626 08/28 02:15:59": "Software Download\\eeprom_save\\2259346202626_991048_2022_08_28 02_15_59\\PSZ26351X35_SOFTWAREDOWNLOAD(991048)_ON_C1000-8FP-2G-L_V1.01_NG(QSFT15).txt"}, "error_count_data": {"2259346202626 08/28 02:15:59": "Software Download\\eeprom_save\\2259346202626_991048_2022_08_28 02_15_59\\PSZ26351X35_SOFTWAREDOWNLOAD(991048)_ON_C1000-8FP-2G-L_V1.01_NG(QSFT15).txt"}, "error_count_path": "Software Download\\eeprom_save"}, 
							"Copy Flash File to BS": {"error_code": "QSFT33", "error_count": "1", "fail_num": "0", "retest_pass_num": "1", "fail_list": {}, "retest_list": {"2259346202255 08/28 00:27:08": "Software Download\\Copy_Flash_File_to_BS\\2259346202255_991072_2022_08_28 00_27_08\\PSZ26351WAQ_SOFTWAREDOWNLOAD(991072)_ON_C1000-8FP-2G-L_V1.01_NG(QSFT33).txt"}, "error_count_data": {"2259346202255 08/28 00:27:08": "Software Download\\Copy_Flash_File_to_BS\\2259346202255_991072_2022_08_28 00_27_08\\PSZ26351WAQ_SOFTWAREDOWNLOAD(991072)_ON_C1000-8FP-2G-L_V1.01_NG(QSFT33).txt"}, "error_count_path": "Software Download\\Copy_Flash_File_to_BS"}, 
							"Boot Loader Check Version": {"error_code": "QSFF69", "error_count": "1", "fail_num": "0", "retest_pass_num": "1", "fail_list": {}, "retest_list": {"2259346200065 08/28 01:16:08": "Software Download\\Boot_Loader_Check_Version\\2259346200065_991036_2022_08_28 01_16_08\\PSZ26351X2P_SOFTWAREDOWNLOAD(991036)_ON_C1000-8FP-2G-L_V1.01_NG(QSFF69).txt"}, "error_count_data": {"2259346200065 08/28 01:16:08": "Software Download\\Boot_Loader_Check_Version\\2259346200065_991036_2022_08_28 01_16_08\\PSZ26351X2P_SOFTWAREDOWNLOAD(991036)_ON_C1000-8FP-2G-L_V1.01_NG(QSFF69).txt"}, "error_count_path": "Software Download\\Boot_Loader_Check_Version"}, 
							"DUT Reload TIME OUT": {"error_code": "QSFCG4", "error_count": "1", "fail_num": "0", "retest_pass_num": "1", "fail_list": {}, "retest_list": {"2259346202964 08/28 05:06:58": "Software Download\\DUT_Reload_TIME_OUT\\2259346202964_991952_2022_08_28 05_06_58\\PSZ26351XPL_SOFTWAREDOWNLOAD(991952)_ON_C1000-8FP-2G-L_V1.01_NG(QSFCG4).txt"}, "error_count_data": {"2259346202964 08/28 05:06:58": "Software Download\\DUT_Reload_TIME_OUT\\2259346202964_991952_2022_08_28 05_06_58\\PSZ26351XPL_SOFTWAREDOWNLOAD(991952)_ON_C1000-8FP-2G-L_V1.01_NG(QSFCG4).txt"}, "error_count_path": "Software Download\\DUT_Reload_TIME_OUT"}}, 
						"final_fail_num": 0, "final_pass_num": 14}}, 

"iplas_data": {
	"user_all_project": ["SWITCH_CISCO_EZ1KA1", "UC_POLY_MTR", "UC_UNIFY_CP700", "EZ1K_A2_ACT2", "SWITCH_YAMAHA_BLUES", "UC_UNIFY_IPPHONE", "UC_UNIFY_IPPHONE__IPK_OPK", "UC_UNIFY_IPPHONE__CP600E"], 
	"time_option": ["Current Shift", "Today", "This Week", "A Week", "YTD Day Shift", "YTD Night Shift", "Select Manually"], 
	"queryid": "bffa8d2suzwonY9y4"}, 
	
"excel_data": {
	"Pretest": {"max_error_count": 2, "max_fail_num": 2, "max_retest_pass_num": 1}, 
	"Hi_pot_test": {"max_error_count": 0, "max_fail_num": 0, "max_retest_pass_num": 0}, 
	"Final_Test": {"max_error_count": 2, "max_fail_num": 2, "max_retest_pass_num": 1}, 
	"Final_Check": {"max_error_count": 1, "max_fail_num": 0, "max_retest_pass_num": 1}, 
	"Software_Download": {"max_error_count": 4, "max_fail_num": 0, "max_retest_pass_num": 4}}} '''

''' save_path = fr"{all_data['today_download_path']}\test.xlsx" '''

cell_style = {'cell_color' : {'spring_green': PatternFill(fill_type="solid", fgColor= "00CCCC"),
                                'yellow': PatternFill(fill_type="solid", fgColor= "FFE5CC"),
                                'light_green': PatternFill(fill_type="solid", fgColor= "a9d08e"),
                                'red': PatternFill(fill_type="solid", fgColor= "e6b8b7"),
                                'retest': PatternFill(fill_type="solid", fgColor= "fde9d9"),
                                'fail': PatternFill(fill_type="solid", fgColor= "dce6f1")},
            'cell_border' : {'all' : lambda line: Border(left=Side(style=line, color='330000'), right=Side(style=line, color='330000'),top=Side(style=line, color='330000'), bottom=Side(style=line, color='330000')), 
                            'btm': lambda line: Border(bottom=Side(style=line, color='330000')),
                            'three_side': lambda line: Border(left=Side(style=line, color='330000'), right=Side(style=line, color='330000'),bottom=Side(style=line, color='330000')),
                            'left': lambda line: Border(left=Side(style=line, color='330000')),
                            } ,
            'cell_font' : {'blod': Font(name = "Calibri", size = 12, bold = True, ),
                            'normal': Font(name = "Calibri", size = 11, ), 
                            'isn': Font(name = "Calibri", size = 11, color='0000FF', ),},
            'cell_alignment': Alignment(horizontal = 'center', vertical = 'center')}
one_page_flag = False

def transfer_to_first_page_data(all_data, execute_data):
    '''
    將all_data['station_data']裡面的資料轉成tuple list形式，並且將執行參數寫成第一頁要的格式
    '''
    global one_page_flag
    frist_page_data = {}
    frist_page_data['iplas_data'] = []
    station_data = all_data['station_data']
    frist_page_data['execute_data'] = [f"Site : {execute_data['site']}", 
                                        f"Project : {execute_data['user_select_project']}", 
                                        f"Data Source : {execute_data['data_source']}",
                                        f"Time : {execute_data['time_selection']['time_period'][0]}~{execute_data['time_selection']['time_period'][1]}"]

    title = ('Station', 'UPH', 'Yield Rate', 'Retest Rate', 'Pass Count', 'Fail Count', 'DUT Retest Pass Count', 'DUT Final Fail Count')
    frist_page_data['iplas_data'].append(title)
    for station_name, station_all_data in station_data.items():
        tmp_dic = []
        tmp_dic.append(station_name)
        tmp_dic.append(station_all_data['UPH'])
        if station_all_data['pass_num']:
            tmp_dic.append(np.round(((station_all_data['pass_num']-station_all_data['final_fail_num'])/station_all_data['pass_num']), decimals = 4))
            tmp_dic.append(np.round((station_all_data['final_pass_num']/station_all_data['pass_num']), decimals = 4))
        else:
            tmp_dic.append(0)
            tmp_dic.append(0)
        tmp_dic.append(station_all_data['pass_num'])
        tmp_dic.append(station_all_data['fail_num'])
        if 'final_pass_num' in station_all_data:
            tmp_dic.append(station_all_data['final_pass_num'])
            tmp_dic.append(station_all_data['final_fail_num'])
        else:
            one_page_flag = True
        frist_page_data['iplas_data'].append(tuple(tmp_dic))
    return frist_page_data

def excel_summary_page(sheet, frist_page_data):
    '''
    填入第一頁summary的資料
    '''
    sheet.title = 'Summary'

    for data in frist_page_data['iplas_data']:  #加入iplas data數據 從第一列開始
        sheet.append(data)                      

    retest_rate_dic = dict()

    #設定iplas data表格的顏色，字型、字體加粗、邊框
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col = sheet.max_column): 
        for cell in row:
            if row[0].row == 1:  #設定第一列title的顏色、字型
                cell.font = cell_style["cell_font"]['blod']
                cell.fill = cell_style["cell_color"]['spring_green']
                cell.alignment = cell_style["cell_alignment"] 
            cell.border = cell_style["cell_border"]['all']('thin')

            if row[0].row > 1 and (cell.column_letter == 'C' or cell.column_letter == 'D'):  #將fail rate、retest rate轉成%格式
                if cell.column_letter == 'D':
                    retest_rate_dic[cell.value] = cell  #將retest rate以dict形式存起來，方便找尋最大值
                cell.number_format = '0.00%'

    #找尋retest rate最大值，並將其填成紅色
    if not one_page_flag:
        value_list = list(retest_rate_dic.keys())
        max_value_index = value_list.index(max(value_list))
        max_cell = retest_rate_dic[max(value_list)]
        max_cell.fill = cell_style["cell_color"]['red']

    start_coordi = (1, 1) #給定表格起始範圍
    end_coordi = (len(frist_page_data['iplas_data']), len(frist_page_data['iplas_data'][0]))  #給定表格結束範圍 
    adjust_sheet_width(sheet, start_coordi, end_coordi)

    #設定第一行的頁超連結和顏色(放這邊是為了防止格式跑掉)
    if not one_page_flag:
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col = sheet.max_column): 
            for cell in row:
                if row[0].row > 1 and cell.column_letter == 'A':    
                    cell.value = '=HYPERLINK("{}", "{}")'.format(f"#{cell.value}!A1", cell.value) 
                    cell.font = cell_style["cell_font"]['isn']

    sheet.insert_rows(1, len(frist_page_data['execute_data'])+1)    #在上方插入5格(表格往下移五格)
    for idx, data in enumerate(frist_page_data['execute_data'], 1):  #填入執行參數的數據
        cell = sheet.cell(row = idx, column=4)
        cell.border = cell_style["cell_border"]['all']('thin')
        sheet.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=4)
        cell = sheet.cell(row = idx, column=1)
        cell.value = data
        cell.fill = cell_style["cell_color"]['yellow']
        cell.font = cell_style["cell_font"]['blod']
        cell.alignment = cell_style["cell_alignment"]

    start_of_draw_data_row = len(frist_page_data['execute_data']) + 2  #開始為執行參數下面兩列
    end_of_draw_data_row = start_of_draw_data_row + len(frist_page_data['iplas_data']) - 1 #結束為加上iplas data的長度
    if not one_page_flag:
        summary_page_draw_barchart(sheet, start_of_draw_data_row, end_of_draw_data_row, max_value_index)   

    #設定所有列高度
    for row in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 18 


def summary_page_draw_barchart(sheet, start_of_draw_data_row, end_of_draw_data_row, max_retest_index):
    chart = BarChart()
    chart.type = 'bar'
    data = Reference(sheet, min_col=4, min_row=start_of_draw_data_row, max_col=4, max_row=end_of_draw_data_row)  #資料拿取: retest rate行數為第四行
    categs = Reference(sheet, min_col=1, min_row=start_of_draw_data_row+1, max_row=end_of_draw_data_row) #y軸拿取title為第一行
    chart.add_data(data, titles_from_data = True)
    chart.set_categories(categs)
    chart.legend = None  #關掉圖例
    #chart.legend.position = 'b' 
    chart.y_axis.majorGridlines = None 
    chart.y_axis.majorUnit = 0.005
    chart.style = 7
    #chart.varyColors = True
    chart.dataLabels = DataLabelList() #顯示數值在直條上面
    chart.dataLabels.showVal = True

    ser1 = chart.series[0]      #把retest rate最高的測站改成紅色
    pt = DataPoint(idx=max_retest_index)
    pt.graphicalProperties.solidFill = "e6b8b7"
    ser1.dPt.append(pt)

    sheet.add_chart(chart, f"A{end_of_draw_data_row+3}") 


def excel_station_page(book, all_data):
    station_data = all_data['station_data']
    max_isn_col = 3
    for station_name, all_station_data in station_data.items():
        #如果error 數目大於3個就已3當作最大，排版用，一個欄位最多就填入3行的isn
        max_error_count = max_isn_col if all_data["excel_data"][station_name]["max_error_count"] > max_isn_col else all_data["excel_data"][station_name]["max_error_count"]  
        max_fail_num = max_isn_col if all_data["excel_data"][station_name]["max_fail_num"] > max_isn_col else all_data["excel_data"][station_name]["max_fail_num"]
        max_retest_pass_num = max_isn_col if all_data["excel_data"][station_name]["max_retest_pass_num"] > max_isn_col else all_data["excel_data"][station_name]["max_retest_pass_num"]

        ISN_data = all_station_data["ISN_data"]
        if not ISN_data:
            continue
        sheet_name = station_name            #將測站名稱有底線的部分以空白替換
        sheet = book.create_sheet(sheet_name)

        #將數據的barchart畫出
        fill_in_barchart_data_and_draw(sheet, ISN_data)

        start_cell = (15, 1) # 除去title列，起始的 row 列, col 行
        now_cell = start_cell
        skip_col = list()           #用於調整寬度的func用，用於跳過isn link那幾行
        need_merge_col = list()     #假設isn link太多造成不只一列，其他行需要合併時，記下需要合併的行

        #填入title名稱 和 調整格式
        sheet.cell(row = now_cell[0]-1, column = now_cell[1]).value = 'error name'
        need_merge_col.append(now_cell[1])
        title_data = (("error count", 'isn link', max_error_count), 
                        ("final fail", 'isn link', max_fail_num), 
                        ("retest pass", 'isn link',max_retest_pass_num))

        #將title的資訊依序填入
        for i in title_data:
            sheet.cell(row = now_cell[0]-1, column = now_cell[1] + 1).value = i[0]
            need_merge_col.append(now_cell[1] + 1)
            sheet.cell(row = now_cell[0]-1, column = now_cell[1] + 2).value = i[1]
            if i[2]:  
                sheet.merge_cells(start_row=now_cell[0]-1, start_column=now_cell[1] + 2, end_row=now_cell[0]-1, end_column=now_cell[1] + i[2] +1) #合併錯誤數title的儲存格
                skip_col.append((get_column_letter(now_cell[1] + 2))) #將合併的起始位置儲存
            
            for cols in range(now_cell[1] + 2, now_cell[1] + i[2] +1 + 1):
                sheet.column_dimensions[get_column_letter(cols)].width = 32  #設置isn link的寬度
            
            now_cell = (now_cell[0], now_cell[1]+ i[2] + 1)
        
        end_col = now_cell[1]
        
        #設定title列的 顏色，字型、邊框
        for col in sheet.iter_rows(min_row=now_cell[0]-1, max_row=now_cell[0]-1, min_col=1, max_col = end_col): 
            for cell in col:
                cell.font = cell_style["cell_font"]['blod']
                cell.alignment = cell_style["cell_alignment"]
                cell.border = cell_style["cell_border"]['all']('medium')
                cell.fill = cell_style["cell_color"]['light_green']

        now_cell = start_cell
        #開始填寫各測站和數據
        for error_name, item_data in ISN_data.items():
            max_height = 0
            sheet.cell(row = now_cell[0], column = now_cell[1]).value = error_name  #填入error name
            fail_data = list(item_data['fail_list'].keys())
            retest_data = list(item_data['retest_list'].keys())

            max_num_data = (("error_count","error_count_data",max_error_count), 
                            ("fail_num","fail_list",max_fail_num), 
                            ("retest_pass_num", "retest_list",max_retest_pass_num))

            #設定三種data 的 數目以及 isn link
            for i in max_num_data:
                sheet.cell(row = now_cell[0], column = now_cell[1] + 1).value = int(item_data[i[0]]) 
                if i[0] == "error_count":  #填入error count的超連結
                    sheet.cell(row=now_cell[0], column=now_cell[1] + 1).value = '=HYPERLINK("{}", "{}")'.format(fr"{all_data['today_download_path']}\{item_data['error_count_path']}", item_data[i[0]])
                height = isn_link_and_max_height(sheet,  all_data, i[1], now_cell[0], now_cell[1] + 2, item_data, fail_data, retest_data)
                now_cell = (now_cell[0], now_cell[1]+ i[2] + 1)
                max_height = height if height > max_height else max_height
        
            if max_height > 1:
                for j in need_merge_col:
                    sheet.merge_cells(start_row=now_cell[0], start_column=j, end_row=now_cell[0]+max_height-1, end_column=j) #合併除了isn link其他行

            
            now_cell = (now_cell[0] + max_height, start_cell[1])
            for col in sheet.iter_rows(min_row=now_cell[0]-1, max_row=now_cell[0]-1, min_col=1, max_col = sheet.max_column): #設定每一列上面有一橫槓
                for cell in col:
                    cell.border = cell_style["cell_border"]['btm']('medium')
        
        start_coordi = (start_cell[0]-1, start_cell[1])  
        end_coordi = (now_cell[0]-1 , end_col)
        adjust_sheet_width(sheet, start_coordi, end_coordi, skip_col)

        for col in sheet.iter_rows(min_row=start_cell[0], max_row=sheet.max_row, min_col=sheet.max_column+1, max_col = sheet.max_column+1): #最右邊的邊框
            for cell in col:
                cell.border = cell_style["cell_border"]['left']('medium')
    
    
def fill_in_barchart_data_and_draw(sheet, ISN_data, max_chart_num = 5):
    '''
    將測項、錯誤總數從A1開始填，並畫barchart在A1上面蓋過數據
    '''
    test_name = list(ISN_data.keys())
    test_data = list(ISN_data.values())
    max_num = max_chart_num if len(test_name) > max_chart_num else len(test_name)  #畫前五多的數據
    sheet.append(('Error Name', 'Error Count'))
    for i in range(max_num, 0 ,-1):
        sheet.append((test_name[i-1], int(test_data[i-1]["error_count"])))

    station_page_draw_barchart(sheet, 1, max_num+1)


def station_page_draw_barchart(sheet, start_of_draw_data_row, end_of_draw_data_row):
    chart = BarChart()
    chart.type = 'bar'
    data = Reference(sheet, min_col=2, min_row=start_of_draw_data_row, max_col=2, max_row=end_of_draw_data_row) # error num col = 2
    categs = Reference(sheet, min_col=1, min_row=start_of_draw_data_row+1, max_row=end_of_draw_data_row)
    chart.add_data(data, titles_from_data = True)
    chart.set_categories(categs)
    chart.legend = None
    chart.y_axis.majorGridlines = None
    chart.y_axis.majorUnit = 1
    chart.varyColors = True
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.gapWidth = 60
    chart.height = 8
    chart.width = 20

    font_test = openpyxl.drawing.text.Font(typeface='Calibri')
    cp = CharacterProperties(latin=font_test, sz=1200)
    chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

    sheet.add_chart(chart, 'A1') 

def isn_link_and_max_height(sheet, all_data, data_type ,now_row, now_column, item_data, fail_data, retest_data):
    '''
    填入isn超連結和找到isn最大高度
    '''
    data_dic = item_data[data_type]
    count = 0
    if len(data_dic):
        for isn, link in data_dic.items():
            sheet.cell(row=now_row, column=now_column).value = '=HYPERLINK("{}", "{}")'.format(fr"{all_data['today_download_path']}\{link}", isn)
            if isn in fail_data:
                sheet.cell(row=now_row, column=now_column).fill = cell_style["cell_color"]['fail']
            if isn in retest_data:
                sheet.cell(row=now_row, column=now_column).fill = cell_style["cell_color"]['retest']
            sheet.cell(row=now_row, column=now_column).font = cell_style["cell_font"]['isn'] 
            sheet.cell(row=now_row, column=now_column).alignment = cell_style["cell_alignment"]
            count += 1    #計算目前填到第幾個
            now_column += 1
            if count % 3 == 0:  #如果填到3的倍數，那就要換下一列
                now_column -= 3
                now_row += 1

        max_height = (len(data_dic) // 3) + 1  #每一列最多3行，多的就換下一列
        if len(data_dic) % 3 == 0:
            max_height = (len(data_dic) // 3)
        return max_height
    return 0
        
def adjust_sheet_width(sheet, start_coordinate, end_coordinate, skip_col = None):
    '''
    給定表格範圍，調整所有格子到適合的大小，
    如果有需要跳過的列，那就調整除了title和合併欄之外的所有字型和置中，並且設定error count那行的寬度以及字體顏色
    '''
    for col in sheet.iter_cols(min_row=start_coordinate[0], max_row=end_coordinate[0], min_col=start_coordinate[1], max_col = end_coordinate[1]):
        if skip_col:
            try:
                letter = col[0].column_letter  #得到該行的代表英文
            except AttributeError as ex:            #假設有except發生，代表是合併儲存格
                if "'MergedCell' object has no attribute 'column_letter'" in str(ex):
                    continue
            if letter in skip_col:
                continue

            for cell in col:
                if cell.row > start_coordinate[0]:
                    cell.border = cell_style["cell_border"]['three_side']('medium')

            if col[0].column_letter == 'B':
                sheet.column_dimensions[col[0].column_letter].width = 16
                for cell in col:
                    if cell.row > start_coordinate[0]:
                        cell.font = cell_style["cell_font"]['isn'] 
                        cell.alignment = cell_style["cell_alignment"]
                continue
          
        for cell in col:
            if cell.row > start_coordinate[0]:
                cell.font = cell_style["cell_font"]['normal']
                cell.alignment = cell_style["cell_alignment"]
                       
        len_list= [len(str(cell.value)) for cell in col if cell.value]
        max_length = max(len_list)
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[col[0].column_letter].width = adjusted_width

    if skip_col:
        for row in range(1, sheet.max_row+1):
            sheet.row_dimensions[row].height = 18

def save_excel(book, today_download_path):
    save_path = fr"{today_download_path}\Summary Report.xlsx"
    book.save(save_path)
    return save_path



def excel_wrtting_flow(all_data, execute_data):
    frist_page_data = transfer_to_first_page_data(all_data, execute_data)
    book = Workbook()
    sheet = book.active
    excel_summary_page(sheet, frist_page_data)
    if not one_page_flag:
        excel_station_page(book, all_data)
    today_download_path = all_data["today_download_path"]
    save_path = save_excel(book, today_download_path)
    return save_path



if '__main__' == __name__:
    excel_wrtting_flow(all_data, execute_data)
    #print(return_to_excel_data(Station_Data))
    #draw_on_excel(Station_Data, User_select_project, Time_period, today_excute_download_path)




   









        
            
        
        



